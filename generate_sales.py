import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import string
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Определяем типы данных для столбцов
data_types = {
    "Дата продажи": "datetime",
    "ID товара": "int",
    "Название товара": "str",
    "Категория товара": "str",
    "Бренд": "str",
    "Цена за единицу": "float",
    "Количество": "int",
    "Сумма продажи": "float",
    "Скидка (%)": "float",
    "Сумма скидки": "float",
    "Итоговая сумма": "float",
    "ID клиента": "int",
    "Имя клиента": "str",
    "Регион": "str",
    "Город": "str",
    "Метод оплаты": "str",
    "Канал продажи": "str",
    "Тип доставки": "str",
    "Статус доставки": "str",
    "Дата доставки": "datetime",
    "Менеджер по продажам": "str",
    "Отдел продаж": "str",
    "Месяц": "int",
    "Год": "int"
}

# Функция для генерации случайной даты в пределах года
def random_date(start_date):
    return start_date + timedelta(days=random.randint(0, 365))

# Функция для генерации данных
def generate_data(row_count=30):
    data = {}
    for col, dtype in data_types.items():
        if dtype == "datetime":
            data[col] = [random_date(datetime(2023, 1, 1)) for _ in range(row_count)]
        elif dtype == "int":
            data[col] = np.random.randint(1, 1000, row_count).tolist()
        elif dtype == "float":
            data[col] = np.round(np.random.uniform(1.0, 1000.0, row_count), 2).tolist()
        elif dtype == "str":
            data[col] = [f"{col}_{random.randint(1, 100)}" for _ in range(row_count)]
    return pd.DataFrame(data)

# Генерация Excel файла с 35 листами
excel_path = "random_sales_data_with_tables_and_styles.xlsx"
with pd.ExcelWriter(excel_path) as writer:
    for i in range(335):
        sheet_data = generate_data()
        sheet_data.to_excel(writer, sheet_name=f"Sheet_{i}", index=False)

# Добавление объектов таблиц и стилей в каждой вкладке
workbook = load_workbook(excel_path)

# Функция для создания случайного имени стиля из 30 букв
def random_style_name():
    return ''.join(random.choices(string.ascii_letters, k=42))

# Создаем 500 уникальных именованных стилей и сохраняем их в список
style_names = []
for i in range(10):
    style_name = random_style_name()
    style = NamedStyle(name=style_name)
    style.font = Font(size=10 + (i % 5), bold=(i % 2 == 0))
    style.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
    workbook.add_named_style(style)
    style_names.append(style_name)

# Функция для генерации случайного цвета в формате RGB
def random_color():
    return f"{random.randint(200, 255):02X}{random.randint(200, 255):02X}{random.randint(200, 255):02X}"

# Генерация списка из 100 уникальных цветов
unique_colors = [random_color() for _ in range(10)]

# Применение таблиц, стилей и ограниченного набора уникальных цветов к каждой вкладке
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]

    end_column = worksheet.max_column
    end_row = worksheet.max_row
    table_ref = f"A1:{chr(64 + end_column)}{end_row}"

    # Создаем объект таблицы с именем вкладки
    table = Table(displayName=sheet_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Применяем стили и уникальные цвета к ячейкам, выбирая случайный цвет из 100 уникальных
    for row in range(1, end_row + 1):
        for col in range(1, end_column + 1):
            # Случайный стиль из созданных
            style_name = random.choice(style_names)
            cell = worksheet.cell(row=row, column=col)
            cell.style = style_name

            # Устанавливаем уникальный цвет для каждой ячейки, выбирая из 100 уникальных цветов
            color = random.choice(unique_colors)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row in range(2, worksheet.max_row + 1):  # начиная со второй строки, так как первая строка - заголовок
        cell = worksheet.cell(row=row, column=1)  # Столбец "A", где "Дата продажи"
        if isinstance(cell.value, datetime):  # Проверяем, что значение является датой
            cell.number_format = 'DD.MM.YYYY'

# Сохраняем изменения
workbook.save(excel_path)
